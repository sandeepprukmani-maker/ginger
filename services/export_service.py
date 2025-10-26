import io
import json
import logging
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExportService:
    def __init__(self):
        pass
    
    def export_to_json(self, stories: List[Dict[str, Any]], epic_data: Dict[str, Any] = None) -> str:
        """
        Export stories to JSON format
        
        Args:
            stories: List of user stories to export
            epic_data: Optional EPIC metadata
            
        Returns:
            JSON string of formatted export data
        """
        try:
            export_data = {
                'exported_at': datetime.now().isoformat(),
                'epic': {
                    'key': epic_data.get('key', '') if epic_data else '',
                    'title': epic_data.get('title', '') if epic_data else '',
                    'description': epic_data.get('description', '') if epic_data else ''
                } if epic_data else None,
                'stories_count': len(stories),
                'stories': stories
            }
            
            return json.dumps(export_data, indent=2, ensure_ascii=False)
            
        except Exception as e:
            logger.error(f"Error exporting to JSON: {str(e)}", exc_info=True)
            raise Exception(f"Failed to export to JSON: {str(e)}")
    
    def export_to_excel(self, stories: List[Dict[str, Any]], epic_data: Dict[str, Any] = None) -> bytes:
        """
        Export stories to Excel (.xlsx) format
        
        Args:
            stories: List of user stories to export
            epic_data: Optional EPIC metadata
            
        Returns:
            Excel file content as bytes
        """
        try:
            wb = Workbook()
            
            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            self._create_summary_sheet(ws_summary, stories, epic_data)
            
            # Stories sheet
            ws_stories = wb.create_sheet("User Stories")
            self._create_stories_sheet(ws_stories, stories)
            
            # Save to bytes
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return excel_file.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}", exc_info=True)
            raise Exception(f"Failed to export to Excel: {str(e)}")
    
    def _create_summary_sheet(self, ws, stories: List[Dict[str, Any]], epic_data: Dict[str, Any] = None):
        """Create summary sheet with EPIC and statistics information"""
        # Header styling
        header_fill = PatternFill(start_color="6B46C1", end_color="6B46C1", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        
        # Title
        ws['A1'] = 'iDraft - User Stories Export'
        ws['A1'].font = Font(bold=True, size=16, color="6B46C1")
        ws.merge_cells('A1:B1')
        
        # Export metadata
        row = 3
        ws[f'A{row}'] = 'Exported At:'
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'A{row}'].font = Font(bold=True)
        
        # EPIC information if available
        if epic_data:
            row += 2
            ws[f'A{row}'] = 'EPIC Information'
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:B{row}')
            
            row += 1
            ws[f'A{row}'] = 'EPIC Key:'
            ws[f'B{row}'] = epic_data.get('key', '')
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = 'Title:'
            ws[f'B{row}'] = epic_data.get('title', '')
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = 'Description:'
            ws[f'B{row}'] = epic_data.get('description', '')[:500]  # Truncate long descriptions
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        
        # Statistics
        row += 2
        ws[f'A{row}'] = 'Statistics'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = 'Total Stories:'
        ws[f'B{row}'] = len(stories)
        ws[f'A{row}'].font = Font(bold=True)
        
        # Priority breakdown
        priority_counts = {}
        for story in stories:
            priority = story.get('priority', 'medium')
            priority_counts[priority] = priority_counts.get(priority, 0) + 1
        
        for priority, count in priority_counts.items():
            row += 1
            ws[f'A{row}'] = f'{priority.capitalize()} Priority:'
            ws[f'B{row}'] = count
        
        # Total story points
        total_points = sum(story.get('story_points', 0) for story in stories)
        row += 1
        ws[f'A{row}'] = 'Total Story Points:'
        ws[f'B{row}'] = total_points
        ws[f'A{row}'].font = Font(bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60
    
    def _create_stories_sheet(self, ws, stories: List[Dict[str, Any]]):
        """Create detailed stories sheet with all story information"""
        # Header row
        headers = ['#', 'Title', 'Description', 'Priority', 'Story Points', 
                   'Developer Criteria', 'QA Criteria', 'Reasoning']
        
        # Styling
        header_fill = PatternFill(start_color="6B46C1", end_color="6B46C1", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write story data
        for row_num, story in enumerate(stories, 2):
            # Story number
            ws.cell(row=row_num, column=1, value=row_num - 1).border = border
            
            # Title
            cell = ws.cell(row=row_num, column=2, value=story.get('title', ''))
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Description
            cell = ws.cell(row=row_num, column=3, value=story.get('description', ''))
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Priority
            priority = story.get('priority', 'medium').upper()
            cell = ws.cell(row=row_num, column=4, value=priority)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            
            # Color code priorities
            if priority == 'HIGH':
                cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            elif priority == 'MEDIUM':
                cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            
            # Story Points
            cell = ws.cell(row=row_num, column=5, value=story.get('story_points', ''))
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            
            # Developer Criteria
            dev_criteria = story.get('developer_criteria', [])
            dev_criteria_text = '\n'.join([f"• {criterion}" for criterion in dev_criteria])
            cell = ws.cell(row=row_num, column=6, value=dev_criteria_text)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # QA Criteria
            qa_criteria = story.get('qa_criteria', [])
            qa_criteria_text = '\n'.join([f"• {criterion}" for criterion in qa_criteria])
            cell = ws.cell(row=row_num, column=7, value=qa_criteria_text)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Reasoning
            cell = ws.cell(row=row_num, column=8, value=story.get('reasoning', ''))
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 40
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def format_stories_as_text(self, stories: List[Dict[str, Any]]) -> str:
        """
        Format stories as readable plain text
        
        Args:
            stories: List of user stories to format
            
        Returns:
            Formatted text string
        """
        formatted_parts = []
        
        for i, story in enumerate(stories, 1):
            formatted_parts.append(f"\n{'='*80}")
            formatted_parts.append(f"USER STORY #{i}: {story.get('title', 'Untitled')}")
            formatted_parts.append(f"{'='*80}\n")
            
            formatted_parts.append(f"Description:\n{story.get('description', '')}\n")
            
            formatted_parts.append(f"Priority: {story.get('priority', 'medium').upper()}")
            formatted_parts.append(f"Story Points: {story.get('story_points', 'TBD')}\n")
            
            if story.get('reasoning'):
                formatted_parts.append(f"Reasoning:\n{story['reasoning']}\n")
            
            formatted_parts.append("Developer Acceptance Criteria:")
            for criterion in story.get('developer_criteria', []):
                formatted_parts.append(f"  • {criterion}")
            
            formatted_parts.append("\nQA Acceptance Criteria:")
            for criterion in story.get('qa_criteria', []):
                formatted_parts.append(f"  • {criterion}")
            
            formatted_parts.append("")
        
        return '\n'.join(formatted_parts)
